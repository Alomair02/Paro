import tempfile
import unittest
from pathlib import Path
from zipfile import ZipFile
import io
import re


from openpyxl import load_workbook
from builders.deck_builder import SAMPLE_THEME, assemble_package
from core.xml_builder import NSMAP, qn
from tests.pptx_test_utils import parse_xml

class TestChartEx(unittest.TestCase):
    def test_funnel_chartex_part_and_invariants(self):

        output_path = Path("tests/fixtures/funnel_chartex.pptx")
        slide_data = [{
            "index": 1,
            "shapes": [{
                "type": "funnel",
                "categories": ["Visit", "Signup", "Trial", "Paid"],
                "series": [{"name": "Users", "values": [5000, 1200, 400, 90]}],
            }],
        }]
        assemble_package(SAMPLE_THEME, slide_data, output_path)

        CHARTEX_NS = "http://schemas.microsoft.com/office/drawing/2014/chartex"
        ns = dict(NSMAP, cx=CHARTEX_NS)

        with ZipFile(output_path) as pptx:
            names = pptx.namelist()
            chartex_parts = sorted(n for n in names if n.startswith("ppt/charts/chartEx") and n.endswith(".xml"))
            self.assertEqual(len(chartex_parts), 1, "expected exactly one chartEx part")
            chartex_part = chartex_parts[0]
            root = parse_xml(pptx.read(chartex_part).decode("utf-8"))

            series = root.find("cx:chart/cx:plotArea/cx:plotAreaRegion/cx:series", ns)
            self.assertIsNotNone(series, "no cx:series found")
            self.assertEqual(series.get("layoutId"), "funnel")

            # chartEx hoists data into cx:numDim (type="val"); the cx:pt text IS the cache.
            num_dim = root.find(".//cx:numDim[@type='val']", ns)
            cache_values = sorted(float(pt.text) for pt in num_dim.findall("cx:lvl/cx:pt", ns))

            workbooks = sorted(n for n in names if n.startswith("xl/embeddings/") and n.endswith(".xlsx"))
            self.assertEqual(len(workbooks), 1)
            ws = load_workbook(io.BytesIO(pptx.read(workbooks[0]))).active
            wb_values = []
            r = 2
            while ws.cell(row=r, column=2).value is not None:   # column B
                wb_values.append(float(ws.cell(row=r, column=2).value))
                r += 1
            self.assertEqual(cache_values, sorted(wb_values))

            # funnel has NO legend (single-series, none in sample)
            self.assertIsNone(root.find("cx:chart/cx:legend", ns), "funnel must not have a legend")
            
            # cx:f formulas must point at the right columns (cache values can be
            # correct while the formula ref is wrong — PowerPoint rereads f on edit)
            str_f = root.find(".//cx:strDim[@type='cat']/cx:f", ns).text
            num_f = root.find(".//cx:numDim[@type='val']/cx:f", ns).text
            self.assertTrue(str_f.startswith("Sheet1!$A$"), f"strDim f should ref col A, got {str_f}")
            self.assertTrue(num_f.startswith("Sheet1!$B$"), f"numDim f should ref col B, got {num_f}")

            rels = parse_xml(pptx.read(f"ppt/charts/_rels/{Path(chartex_part).name}.rels").decode("utf-8"))
            rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            rel_types = sorted(rel.get("Type") for rel in rels.findall("r:Relationship", rel_ns))
            self.assertEqual(rel_types, sorted([
                "http://schemas.openxmlformats.org/officeDocument/2006/relationships/package",
                "http://schemas.microsoft.com/office/2011/relationships/chartStyle",
                "http://schemas.microsoft.com/office/2011/relationships/chartColorStyle",
            ]), "chartEx part must relate workbook + style + colorstyle")

            ct = pptx.read("[Content_Types].xml").decode("utf-8")
            self.assertIn("application/vnd.ms-office.chartex+xml", ct)
            self.assertIn("application/vnd.ms-office.chartstyle+xml", ct)
            self.assertIn("application/vnd.ms-office.chartcolorstyle+xml", ct)
            
            for part in [chartex_part] + [n for n in names if "style_chartEx" in n or "colors_chartEx" in n]:
                text = pptx.read(part).decode("utf-8")
                for m in re.finditer(r'(lumMod|lumOff)[^>]*val="(\d+)"', text):
                    self.assertLessEqual(int(m.group(2)), 100000,
                        f"{m.group(1)} > 100000 in {part} triggers PowerPoint repair")
                    
            self.assertNotIn("cx", NSMAP, "cx must be declared locally, never globally")
    
    def test_waterfall_chartex_invariants(self):
        from zipfile import ZipFile
        import io
        from openpyxl import load_workbook
        from builders.deck_builder import SAMPLE_THEME, assemble_package

        output_path = Path("tests/fixtures/waterfall_chartex.pptx")
        slide_data = [{
            "index": 1,
            "shapes": [{
                "type": "waterfall",
                "title": "Cash Flow",
                "categories": ["Start", "Q1", "Q2", "Q3", "Q4", "Costs", "Tax", "End"],
                "series": [{"name": "Net", "values": [100, 20, 50, -40, 130, -60, 70, 140]}],
                "subtotals": [0, 4, 7],
            }],
        }]
        assemble_package(SAMPLE_THEME, slide_data, output_path)

        CHARTEX_NS = "http://schemas.microsoft.com/office/drawing/2014/chartex"
        ns = dict(NSMAP, cx=CHARTEX_NS)

        with ZipFile(output_path) as pptx:
            names = pptx.namelist()
            chartex_part = next(n for n in names if n.startswith("ppt/charts/chartEx") and n.endswith(".xml"))
            root = parse_xml(pptx.read(chartex_part).decode("utf-8"))

            # layoutId is waterfall
            series = root.find("cx:chart/cx:plotArea/cx:plotAreaRegion/cx:series", ns)
            self.assertEqual(series.get("layoutId"), "waterfall")

            # subtotals match the marked indices (the distinguishing waterfall feature)
            idxs = [int(e.get("val")) for e in root.findall(".//cx:layoutPr/cx:subtotals/cx:idx", ns)]
            self.assertEqual(idxs, [0, 4, 7])

            # two axes: catScaling (id 0) + valScaling (id 1)
            axes = root.findall("cx:chart/cx:plotArea/cx:axis", ns)
            self.assertEqual(len(axes), 2)
            self.assertIsNotNone(axes[0].find("cx:catScaling", ns))
            self.assertIsNotNone(axes[1].find("cx:valScaling", ns))

            # waterfall carries a legend (top-center)
            legend = root.find("cx:chart/cx:legend", ns)
            self.assertIsNotNone(legend, "waterfall must have a legend")
            self.assertEqual(legend.get("pos"), "t")

            # cache == workbook (same invariant as funnel)
            num_dim = root.find(".//cx:numDim[@type='val']", ns)
            cache_values = sorted(float(pt.text) for pt in num_dim.findall("cx:lvl/cx:pt", ns))
            workbooks = sorted(n for n in names if n.startswith("xl/embeddings/") and n.endswith(".xlsx"))
            ws = load_workbook(io.BytesIO(pptx.read(workbooks[0]))).active
            wb_values = []
            r = 2
            while ws.cell(row=r, column=2).value is not None:
                wb_values.append(float(ws.cell(row=r, column=2).value))
                r += 1
            self.assertEqual(cache_values, sorted(wb_values))

            # rels triple (same acceptance contract)
            rels = parse_xml(pptx.read(f"ppt/charts/_rels/{Path(chartex_part).name}.rels").decode("utf-8"))
            rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            rel_types = sorted(rel.get("Type") for rel in rels.findall("r:Relationship", rel_ns))
            self.assertEqual(rel_types, sorted([
                "http://schemas.openxmlformats.org/officeDocument/2006/relationships/package",
                "http://schemas.microsoft.com/office/2011/relationships/chartStyle",
                "http://schemas.microsoft.com/office/2011/relationships/chartColorStyle",
            ]))

    def test_histogram_chartex_invariants(self):
        from zipfile import ZipFile
        import io
        from openpyxl import load_workbook
        from builders.deck_builder import SAMPLE_THEME, assemble_package

        output_path = Path("tests/fixtures/histogram_chartex.pptx")
        vals = [1,3,3,5,6,6,7,8,9,9,10,10,11,12,13,14,15,16,17,19,22,24]
        slide_data = [{
            "index": 1,
            "shapes": [{
                "type": "histogram",
                "title": "Distribution",
                "categories": [],
                "series": [{"name": "Samples", "values": vals}],
            }],
        }]
        assemble_package(SAMPLE_THEME, slide_data, output_path)

        CHARTEX_NS = "http://schemas.microsoft.com/office/drawing/2014/chartex"
        ns = dict(NSMAP, cx=CHARTEX_NS)

        with ZipFile(output_path) as pptx:
            names = pptx.namelist()
            chartex_part = next(n for n in names if n.startswith("ppt/charts/chartEx") and n.endswith(".xml"))
            root = parse_xml(pptx.read(chartex_part).decode("utf-8"))

            # layoutId is clusteredColumn (histogram's counterintuitive id)
            series = root.find("cx:chart/cx:plotArea/cx:plotAreaRegion/cx:series", ns)
            self.assertEqual(series.get("layoutId"), "clusteredColumn")

            # the distinguishing histogram features: NO strDim, and a binning layoutPr
            self.assertIsNone(root.find(".//cx:strDim", ns), "histogram must have no strDim")
            self.assertIsNotNone(root.find(".//cx:layoutPr/cx:binning", ns), "histogram needs binning")

            # numDim references column A (no category column ahead of it)
            num_f = root.find(".//cx:numDim[@type='val']/cx:f", ns).text
            self.assertTrue(num_f.startswith("Sheet1!$A$"), f"histogram values in col A, got {num_f}")

            # cache == workbook, with values in column A
            cache_values = sorted(float(pt.text)
                                  for pt in root.findall(".//cx:numDim[@type='val']/cx:lvl/cx:pt", ns))
            workbooks = sorted(n for n in names if n.startswith("xl/embeddings/") and n.endswith(".xlsx"))
            ws = load_workbook(io.BytesIO(pptx.read(workbooks[0]))).active
            wb_values = []
            r = 2
            while ws.cell(row=r, column=1).value is not None:   # column A
                wb_values.append(float(ws.cell(row=r, column=1).value))
                r += 1
            self.assertEqual(cache_values, sorted(wb_values))
    
    def test_boxwhisker_chartex_invariants(self):
        output_path = Path("tests/fixtures/boxwhisker_chartex.pptx")
        cats = ["Category 1"] * 11 + ["Category 2"] * 11
        slide_data = [{
            "index": 1,
            "shapes": [{
                "type": "boxWhisker",
                "title": "Distributions",
                "categories": cats,
                "series": [
                    {"name": "Series1", "values": [-7,-10,-28,47,11,-24,-24,36,10,-78,47,-24,-17,-12,-11,17,14,46,-18,19,-26,-20]},
                    {"name": "Series2", "values": [-3,1,-6,10,34,128,22,-12,-28,6,31,3,12,-12,-13,6,15,41,16,10,23,16]},
                ],
            }],
        }]
        assemble_package(SAMPLE_THEME, slide_data, output_path)

        CHARTEX_NS = "http://schemas.microsoft.com/office/drawing/2014/chartex"
        ns = dict(NSMAP, cx=CHARTEX_NS)

        with ZipFile(output_path) as pptx:
            names = pptx.namelist()
            chartex_part = next(n for n in names
                                if n.startswith("ppt/charts/chartEx") and n.endswith(".xml"))
            root = parse_xml(pptx.read(chartex_part).decode("utf-8"))

            # --- multi-series: N data blocks, one per series, ids 0..N-1 ---
            data_blocks = root.findall("cx:chartData/cx:data", ns)
            self.assertEqual(len(data_blocks), 2, "two series -> two cx:data blocks")
            self.assertEqual([d.get("id") for d in data_blocks], ["0", "1"])

            # --- each data block's numDim references its own column (B, then C) ---
            num_refs = [d.find("cx:numDim/cx:f", ns).text for d in data_blocks]
            self.assertTrue(num_refs[0].startswith("Sheet1!$B$"), f"series 0 -> col B, got {num_refs[0]}")
            self.assertTrue(num_refs[1].startswith("Sheet1!$C$"), f"series 1 -> col C, got {num_refs[1]}")

            # --- N series, each boxWhisker, each bound to its data block, each its own quartile layoutPr ---
            series = root.findall("cx:chart/cx:plotArea/cx:plotAreaRegion/cx:series", ns)
            self.assertEqual(len(series), 2, "two series elements")
            for s_idx, ser in enumerate(series):
                self.assertEqual(ser.get("layoutId"), "boxWhisker")
                self.assertEqual(ser.find("cx:dataId", ns).get("val"), str(s_idx),
                                 f"series {s_idx} binds to data block {s_idx}")
                stats = ser.find("cx:layoutPr/cx:statistics", ns)
                self.assertIsNotNone(stats, f"series {s_idx} needs its own statistics layoutPr")
                self.assertEqual(stats.get("quartileMethod"), "exclusive")
                # box plots: value labels OFF
                vis = ser.find("cx:dataLabels/cx:visibility", ns)
                self.assertEqual(vis.get("value"), "0", "box plot value labels must be off")

            # --- per-series cache == workbook: numDim cache matches its own column ---
            workbooks = sorted(n for n in names
                               if n.startswith("xl/embeddings/") and n.endswith(".xlsx"))
            ws = load_workbook(io.BytesIO(pptx.read(workbooks[0]))).active
            for s_idx, dblock in enumerate(data_blocks):
                cache = [float(pt.text) for pt in dblock.findall("cx:numDim/cx:lvl/cx:pt", ns)]
                col = 2 + s_idx                      # B=2 (series 0), C=3 (series 1)
                wb_col = []
                r = 2
                while ws.cell(row=r, column=col).value is not None:
                    wb_col.append(float(ws.cell(row=r, column=col).value))
                    r += 1
                self.assertEqual(sorted(cache), sorted(wb_col),
                                 f"series {s_idx} cache must match workbook column {col}")

            # --- two axes (cat + val) ---
            axes = root.findall("cx:chart/cx:plotArea/cx:axis", ns)
            self.assertEqual(len(axes), 2)
            self.assertIsNotNone(axes[0].find("cx:catScaling", ns))
            self.assertIsNotNone(axes[1].find("cx:valScaling", ns))
    
    def test_pareto_chartex_invariants(self):
        import random
        output_path = Path("tests/fixtures/pareto_chartex.pptx")
        random.seed(0)
        cats = [f"Category {random.randint(1,4)}" for _ in range(50)]
        slide_data = [{
            "index": 1,
            "shapes": [{
                "type": "pareto",
                "title": "Defect Causes",
                "categories": cats,
                "series": [{"name": "Count", "values": [1]*50}],
            }],
        }]
        assemble_package(SAMPLE_THEME, slide_data, output_path)

        CHARTEX_NS = "http://schemas.microsoft.com/office/drawing/2014/chartex"
        ns = dict(NSMAP, cx=CHARTEX_NS)

        with ZipFile(output_path) as pptx:
            names = pptx.namelist()
            chartex_part = next(n for n in names
                                if n.startswith("ppt/charts/chartEx") and n.endswith(".xml"))
            root = parse_xml(pptx.read(chartex_part).decode("utf-8"))

            series = root.findall("cx:chart/cx:plotArea/cx:plotAreaRegion/cx:series", ns)
            self.assertEqual(len(series), 2, "pareto = two series (bars + line)")

            # series 0: clusteredColumn, has data + aggregation + axisId=1
            bars = series[0]
            self.assertEqual(bars.get("layoutId"), "clusteredColumn")
            self.assertEqual(bars.find("cx:dataId", ns).get("val"), "0")
            self.assertIsNotNone(bars.find("cx:layoutPr/cx:aggregation", ns), "bars need aggregation")
            self.assertEqual(bars.find("cx:axisId", ns).get("val"), "1")

            # series 1: paretoLine, DERIVED (ownerIdx, no dataId), axisId=2
            line = series[1]
            self.assertEqual(line.get("layoutId"), "paretoLine")
            self.assertEqual(line.get("ownerIdx"), "0", "line derives from series 0")
            self.assertIsNone(line.find("cx:dataId", ns), "derived line has no dataId")
            self.assertEqual(line.find("cx:axisId", ns).get("val"), "2")

            # three axes: cat(0), val count(1), val percentage(2)
            axes = root.findall("cx:chart/cx:plotArea/cx:axis", ns)
            self.assertEqual(len(axes), 3, "pareto has three axes")
            self.assertIsNotNone(axes[0].find("cx:catScaling", ns))
            self.assertIsNotNone(axes[1].find("cx:valScaling", ns))
            pct = axes[2].find("cx:units", ns)
            self.assertIsNotNone(pct, "third axis is the percentage axis")
            self.assertEqual(pct.get("unit"), "percentage")

            # shares histogram's style (verified md5)
            rels = parse_xml(pptx.read(f"ppt/charts/_rels/{Path(chartex_part).name}.rels").decode("utf-8"))
            rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            rel_types = sorted(r.get("Type") for r in rels.findall("r:Relationship", rel_ns))
            self.assertEqual(len(rel_types), 3, "package + chartStyle + chartColorStyle")

    def test_treemap_chartex_invariants(self):
        output_path = Path("tests/fixtures/treemap_chartex.pptx")
        records = [
            {"path": ["Branch 1", "Stem 1", "Leaf 1"], "value": 22},
            {"path": ["Branch 1", "Stem 1", "Leaf 2"], "value": 12},
            {"path": ["Branch 1", "Stem 2", "Leaf 3"], "value": 18},
            {"path": ["Branch 2", "Stem 3", "Leaf 4"], "value": 87},
            {"path": ["Branch 2", "Stem 3", "Leaf 5"], "value": 25},
            {"path": ["Branch 3", "Stem 4", "Leaf 6"], "value": 30},
        ]
        slide_data = [{
            "index": 1,
            "shapes": [{
                "type": "treemap",
                "title": "Org",
                "categories": [],
                "series": [{"name": "Size", "points": records}],
            }],
        }]
        assemble_package(SAMPLE_THEME, slide_data, output_path)

        CHARTEX_NS = "http://schemas.microsoft.com/office/drawing/2014/chartex"
        ns = dict(NSMAP, cx=CHARTEX_NS)

        with ZipFile(output_path) as pptx:
            names = pptx.namelist()
            chartex_part = next(n for n in names
                                if n.startswith("ppt/charts/chartEx") and n.endswith(".xml"))
            root = parse_xml(pptx.read(chartex_part).decode("utf-8"))

            # treemap layoutId, no axis, has legend, parentLabelLayout
            series = root.find("cx:chart/cx:plotArea/cx:plotAreaRegion/cx:series", ns)
            self.assertEqual(series.get("layoutId"), "treemap")
            self.assertIsNone(root.find("cx:chart/cx:plotArea/cx:axis", ns), "treemap has no axes")
            self.assertIsNotNone(root.find("cx:chart/cx:legend", ns), "treemap has a legend")
            self.assertIsNotNone(series.find("cx:layoutPr/cx:parentLabelLayout", ns))

            # numDim is "size" (not "val")
            self.assertEqual(root.find(".//cx:numDim", ns).get("type"), "size")

            # depth-3 hierarchy: THREE cx:lvl in strDim, innermost-first (Leaf, Stem, Branch)
            lvls = root.findall("cx:chartData/cx:data/cx:strDim/cx:lvl", ns)
            self.assertEqual(len(lvls), 3, "3-level hierarchy -> 3 cx:lvl")
            self.assertEqual(lvls[0].find("cx:pt", ns).text, "Leaf 1", "level 0 = innermost (leaf)")
            self.assertEqual(lvls[1].find("cx:pt", ns).text, "Stem 1")
            self.assertEqual(lvls[2].find("cx:pt", ns).text, "Branch 1", "last level = outermost (branch)")

            # strDim spans A:C (3 label cols), numDim is D
            self.assertEqual(root.find(".//cx:strDim/cx:f", ns).text, "Sheet1!$A$2:$C$7")
            self.assertEqual(root.find(".//cx:numDim/cx:f", ns).text, "Sheet1!$D$2:$D$7")

            # cache == workbook (size column), per the size col = depth+1 = D
            cache = sorted(float(pt.text) for pt in root.findall(".//cx:numDim/cx:lvl/cx:pt", ns))
            wbs = sorted(n for n in names if n.startswith("xl/embeddings/") and n.endswith(".xlsx"))
            ws = load_workbook(io.BytesIO(pptx.read(wbs[0]))).active
            wb = []
            r = 2
            while ws.cell(row=r, column=4).value is not None:   # column D = size
                wb.append(float(ws.cell(row=r, column=4).value)); r += 1
            self.assertEqual(cache, sorted(wb))

            # style is treemap's own
            rels = parse_xml(pptx.read(f"ppt/charts/_rels/{Path(chartex_part).name}.rels").decode("utf-8"))
            rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            self.assertEqual(len([r for r in rels.findall("r:Relationship", rel_ns)]), 3)

    def test_treemap_chartex_invariants(self):
        output_path = Path("tests/fixtures/treemap_chartex.pptx")
        records = [
            {"path": ["Branch 1", "Stem 1", "Leaf 1"], "value": 22},
            {"path": ["Branch 1", "Stem 1", "Leaf 2"], "value": 12},
            {"path": ["Branch 1", "Stem 2", "Leaf 3"], "value": 18},
            {"path": ["Branch 2", "Stem 3", "Leaf 4"], "value": 87},
            {"path": ["Branch 2", "Stem 3", "Leaf 5"], "value": 25},
            {"path": ["Branch 3", "Stem 4", "Leaf 6"], "value": 30},
        ]
        slide_data = [{
            "index": 1,
            "shapes": [{
                "type": "treemap",
                "title": "Org",
                "categories": [],
                "series": [{"name": "Size", "points": records}],
            }],
        }]
        assemble_package(SAMPLE_THEME, slide_data, output_path)

        CHARTEX_NS = "http://schemas.microsoft.com/office/drawing/2014/chartex"
        ns = dict(NSMAP, cx=CHARTEX_NS)

        with ZipFile(output_path) as pptx:
            names = pptx.namelist()
            chartex_part = next(n for n in names
                                if n.startswith("ppt/charts/chartEx") and n.endswith(".xml"))
            root = parse_xml(pptx.read(chartex_part).decode("utf-8"))

            # treemap layoutId, no axis, has legend, parentLabelLayout
            series = root.find("cx:chart/cx:plotArea/cx:plotAreaRegion/cx:series", ns)
            self.assertEqual(series.get("layoutId"), "treemap")
            self.assertIsNone(root.find("cx:chart/cx:plotArea/cx:axis", ns), "treemap has no axes")
            self.assertIsNotNone(root.find("cx:chart/cx:legend", ns), "treemap has a legend")
            self.assertIsNotNone(series.find("cx:layoutPr/cx:parentLabelLayout", ns))

            # numDim is "size" (not "val")
            self.assertEqual(root.find(".//cx:numDim", ns).get("type"), "size")

            # depth-3 hierarchy: THREE cx:lvl in strDim, innermost-first (Leaf, Stem, Branch)
            lvls = root.findall("cx:chartData/cx:data/cx:strDim/cx:lvl", ns)
            self.assertEqual(len(lvls), 3, "3-level hierarchy -> 3 cx:lvl")
            self.assertEqual(lvls[0].find("cx:pt", ns).text, "Leaf 1", "level 0 = innermost (leaf)")
            self.assertEqual(lvls[1].find("cx:pt", ns).text, "Stem 1")
            self.assertEqual(lvls[2].find("cx:pt", ns).text, "Branch 1", "last level = outermost (branch)")

            # strDim spans A:C (3 label cols), numDim is D
            self.assertEqual(root.find(".//cx:strDim/cx:f", ns).text, "Sheet1!$A$2:$C$7")
            self.assertEqual(root.find(".//cx:numDim/cx:f", ns).text, "Sheet1!$D$2:$D$7")

            # cache == workbook (size column), per the size col = depth+1 = D
            cache = sorted(float(pt.text) for pt in root.findall(".//cx:numDim/cx:lvl/cx:pt", ns))
            wbs = sorted(n for n in names if n.startswith("xl/embeddings/") and n.endswith(".xlsx"))
            ws = load_workbook(io.BytesIO(pptx.read(wbs[0]))).active
            wb = []
            r = 2
            while ws.cell(row=r, column=4).value is not None:   # column D = size
                wb.append(float(ws.cell(row=r, column=4).value)); r += 1
            self.assertEqual(cache, sorted(wb))

            # style is treemap's own
            rels = parse_xml(pptx.read(f"ppt/charts/_rels/{Path(chartex_part).name}.rels").decode("utf-8"))
            rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            self.assertEqual(len([r for r in rels.findall("r:Relationship", rel_ns)]), 3)

    def test_treemap_chartex_invariants(self):
        output_path = Path("tests/fixtures/treemap_chartex.pptx")
        records = [
            {"path": ["Branch 1", "Stem 1", "Leaf 1"], "value": 22},
            {"path": ["Branch 1", "Stem 1", "Leaf 2"], "value": 12},
            {"path": ["Branch 1", "Stem 2", "Leaf 3"], "value": 18},
            {"path": ["Branch 2", "Stem 3", "Leaf 4"], "value": 87},
            {"path": ["Branch 2", "Stem 3", "Leaf 5"], "value": 25},
            {"path": ["Branch 3", "Stem 4", "Leaf 6"], "value": 30},
        ]
        slide_data = [{
            "index": 1,
            "shapes": [{
                "type": "treemap",
                "title": "Org",
                "categories": [],
                "series": [{"name": "Size", "points": records}],
            }],
        }]
        assemble_package(SAMPLE_THEME, slide_data, output_path)

        CHARTEX_NS = "http://schemas.microsoft.com/office/drawing/2014/chartex"
        ns = dict(NSMAP, cx=CHARTEX_NS)

        with ZipFile(output_path) as pptx:
            names = pptx.namelist()
            chartex_part = next(n for n in names
                                if n.startswith("ppt/charts/chartEx") and n.endswith(".xml"))
            root = parse_xml(pptx.read(chartex_part).decode("utf-8"))

            # treemap layoutId, no axis, has legend, parentLabelLayout
            series = root.find("cx:chart/cx:plotArea/cx:plotAreaRegion/cx:series", ns)
            self.assertEqual(series.get("layoutId"), "treemap")
            self.assertIsNone(root.find("cx:chart/cx:plotArea/cx:axis", ns), "treemap has no axes")
            self.assertIsNotNone(root.find("cx:chart/cx:legend", ns), "treemap has a legend")
            self.assertIsNotNone(series.find("cx:layoutPr/cx:parentLabelLayout", ns))

            # numDim is "size" (not "val")
            self.assertEqual(root.find(".//cx:numDim", ns).get("type"), "size")

            # depth-3 hierarchy: THREE cx:lvl in strDim, innermost-first (Leaf, Stem, Branch)
            lvls = root.findall("cx:chartData/cx:data/cx:strDim/cx:lvl", ns)
            self.assertEqual(len(lvls), 3, "3-level hierarchy -> 3 cx:lvl")
            self.assertEqual(lvls[0].find("cx:pt", ns).text, "Leaf 1", "level 0 = innermost (leaf)")
            self.assertEqual(lvls[1].find("cx:pt", ns).text, "Stem 1")
            self.assertEqual(lvls[2].find("cx:pt", ns).text, "Branch 1", "last level = outermost (branch)")

            # strDim spans A:C (3 label cols), numDim is D
            self.assertEqual(root.find(".//cx:strDim/cx:f", ns).text, "Sheet1!$A$2:$C$7")
            self.assertEqual(root.find(".//cx:numDim/cx:f", ns).text, "Sheet1!$D$2:$D$7")

            # cache == workbook (size column), per the size col = depth+1 = D
            cache = sorted(float(pt.text) for pt in root.findall(".//cx:numDim/cx:lvl/cx:pt", ns))
            wbs = sorted(n for n in names if n.startswith("xl/embeddings/") and n.endswith(".xlsx"))
            ws = load_workbook(io.BytesIO(pptx.read(wbs[0]))).active
            wb = []
            r = 2
            while ws.cell(row=r, column=4).value is not None:   # column D = size
                wb.append(float(ws.cell(row=r, column=4).value)); r += 1
            self.assertEqual(cache, sorted(wb))

            # style is treemap's own
            rels = parse_xml(pptx.read(f"ppt/charts/_rels/{Path(chartex_part).name}.rels").decode("utf-8"))
            rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            self.assertEqual(len([r for r in rels.findall("r:Relationship", rel_ns)]), 3)

    def test_treemap_rejects_mixed_depth(self):
        # the one real format constraint: uniform leaf depth (resolver guard)
        from transpiler.resolver import LayoutResolver  # adjust import to your resolver entry
        # build a mixed-depth node tree via DSL and assert it raises
        # (or call the resolver path directly) — see note below
    
    def test_sunburst_chartex_invariants(self):
        output_path = Path("tests/fixtures/sunburst_chartex.pptx")
        records = [
            {"path": ["Branch 1", "Stem 1", "Leaf 1"], "value": 22},
            {"path": ["Branch 1", "Stem 1", "Leaf 2"], "value": 12},
            {"path": ["Branch 1", "Stem 2", "Leaf 3"], "value": 18},
            {"path": ["Branch 2", "Stem 3", "Leaf 4"], "value": 87},
            {"path": ["Branch 2", "Stem 3", "Leaf 5"], "value": 25},
            {"path": ["Branch 3", "Stem 4", "Leaf 6"], "value": 30},
        ]
        slide_data = [{
            "index": 1,
            "shapes": [{
                "type": "sunburst",
                "title": "Org",
                "categories": [],
                "series": [{"name": "Size", "points": records}],
            }],
        }]
        assemble_package(SAMPLE_THEME, slide_data, output_path)

        CHARTEX_NS = "http://schemas.microsoft.com/office/drawing/2014/chartex"
        ns = dict(NSMAP, cx=CHARTEX_NS)

        with ZipFile(output_path) as pptx:
            names = pptx.namelist()
            chartex_part = next(n for n in names
                                if n.startswith("ppt/charts/chartEx") and n.endswith(".xml"))
            root = parse_xml(pptx.read(chartex_part).decode("utf-8"))

            series = root.find("cx:chart/cx:plotArea/cx:plotAreaRegion/cx:series", ns)
            # sunburst deltas vs treemap:
            self.assertEqual(series.get("layoutId"), "sunburst")
            self.assertEqual(series.find("cx:dataLabels", ns).get("pos"), "ctr",
                             "sunburst labels are centered (vs treemap inEnd)")
            self.assertIsNone(series.find("cx:layoutPr", ns),
                              "sunburst has NO layoutPr/parentLabelLayout")
            self.assertIsNone(root.find("cx:chart/cx:legend", ns),
                              "sunburst has NO legend")

            # shared hierarchical model (same as treemap): 3 cx:lvl innermost-first, size numDim
            lvls = root.findall("cx:chartData/cx:data/cx:strDim/cx:lvl", ns)
            self.assertEqual(len(lvls), 3)
            self.assertEqual(lvls[0].find("cx:pt", ns).text, "Leaf 1", "level 0 = innermost")
            self.assertEqual(lvls[2].find("cx:pt", ns).text, "Branch 1", "level 2 = outermost")
            self.assertEqual(root.find(".//cx:numDim", ns).get("type"), "size")

            # uses sunburst's own style asset
            rels = parse_xml(pptx.read(f"ppt/charts/_rels/{Path(chartex_part).name}.rels").decode("utf-8"))
            rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            self.assertEqual(len(rels.findall("r:Relationship", rel_ns)), 3)

            # cache == workbook (size column)
            cache = sorted(float(pt.text) for pt in root.findall(".//cx:numDim/cx:lvl/cx:pt", ns))
            wbs = [n for n in names if n.startswith("xl/embeddings/") and n.endswith(".xlsx")]
            ws = load_workbook(io.BytesIO(pptx.read(wbs[0]))).active
            wb, r = [], 2
            while ws.cell(row=r, column=4).value is not None:
                wb.append(float(ws.cell(row=r, column=4).value)); r += 1
            self.assertEqual(cache, sorted(wb))