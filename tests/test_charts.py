import unittest
import io
from pathlib import Path
from zipfile import ZipFile
from builders.deck_builder import SAMPLE_THEME, assemble_package
from tests.pptx_test_utils import parse_xml
from openpyxl import load_workbook

class TestCharts(unittest.TestCase):
    def test_radar_chart_invariants(self):
        output_path = Path("tests/fixtures/radar_chart.pptx")
        slide_data = [{
            "index": 1,
            "shapes": [{
                "type": "chart", "chart_type": "radar",
                "title": "This is my radar",
                "categories": ["Mon", "Tue", "Wed", "Thu", "Fri"],
                "series": [
                    {"name": "Series 1", "values": [32, 32, 28, 12, 15]},
                    {"name": "Series 2", "values": [12, 12, 12, 21, 28]},
                ],
            }],
        }]
        assemble_package(SAMPLE_THEME, slide_data, output_path)

        C_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
        ns = {"c": C_NS}

        with ZipFile(output_path) as pptx:
            names = pptx.namelist()
            chart_part = next(n for n in names
                              if n.startswith("ppt/charts/chart") and n.endswith(".xml")
                              and "chartEx" not in n and "_rels" not in n
                              and "colors" not in n and "style" not in n)
            root = parse_xml(pptx.read(chart_part).decode("utf-8"))

            # the radar plot element exists
            radar = root.find(".//c:plotArea/c:radarChart", ns)
            self.assertIsNotNone(radar, "must emit c:radarChart")

            # radarStyle (radar's signature element)
            rs = radar.find("c:radarStyle", ns)
            self.assertIsNotNone(rs, "radar needs radarStyle")
            self.assertEqual(rs.get("val"), "marker")

            # two series, each with a marker symbol=none
            sers = radar.findall("c:ser", ns)
            self.assertEqual(len(sers), 2, "two series")
            for ser in sers:
                sym = ser.find("c:marker/c:symbol", ns)
                self.assertIsNotNone(sym, "radar series needs a marker")
                self.assertEqual(sym.get("val"), "none")
                # radar must NOT carry c:smooth (that's line-only)
                self.assertIsNone(ser.find("c:smooth", ns), "radar series must not have smooth")

            # catAx + valAx pair, crossAx cross-referencing each other
            cat_ax = root.find(".//c:plotArea/c:catAx", ns)
            val_ax = root.find(".//c:plotArea/c:valAx", ns)
            self.assertIsNotNone(cat_ax, "radar has a catAx")
            self.assertIsNotNone(val_ax, "radar has a valAx")
            cat_id = cat_ax.find("c:axId", ns).get("val")
            val_id = val_ax.find("c:axId", ns).get("val")
            self.assertEqual(cat_ax.find("c:crossAx", ns).get("val"), val_id)
            self.assertEqual(val_ax.find("c:crossAx", ns).get("val"), cat_id)

            # THE RINGS: valAx majorGridlines (the radar skeleton — the bug that bit)
            self.assertIsNotNone(val_ax.find("c:majorGridlines", ns),
                                 "radar valAx must have majorGridlines (the concentric rings)")

            # cache == workbook (standard classic-chart invariant)
            cache = []
            for ser in sers:
                cache.extend(float(pt.find("c:v", ns).text)
                             for pt in ser.findall("c:val/c:numRef/c:numCache/c:pt", ns))
            wbs = [n for n in names if n.startswith("xl/embeddings/") and n.endswith(".xlsx")]
            ws = load_workbook(io.BytesIO(pptx.read(wbs[0]))).active
            wb = []
            for col in ("B", "C"):
                r = 2
                while ws[f"{col}{r}"].value is not None:
                    wb.append(float(ws[f"{col}{r}"].value)); r += 1
            self.assertEqual(sorted(cache), sorted(wb), "cache must equal workbook")

class ChartDefaultsTests(unittest.TestCase):
    """Probe #7 guards: zero-baseline value axes and default-on value labels."""

    C_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
    ns = {"c": C_NS}

    @staticmethod
    def _build(chart_type="column", values=(3.8, 4.1, 4.5), **kwargs):
        from builders.chart_part_builder import build_chart_part_xml
        series = kwargs.pop("series_list", None) or [
            {"name": "s", "values": list(values)}
        ]
        xml = build_chart_part_xml(
            "rId1", ["A", "B", "C"], series, chart_type, **kwargs
        )
        return parse_xml(xml)

    def test_column_value_axis_baselines_at_zero(self):
        root = self._build("column", values=(3.8, 4.1, 4.5))
        ax_min = root.find(".//c:valAx/c:scaling/c:min", self.ns)
        self.assertIsNotNone(ax_min, "clustered values must not float the baseline")
        self.assertEqual(ax_min.get("val"), "0")

    def test_column_with_negative_values_keeps_auto_min(self):
        root = self._build("column", values=(-2, 4, 5))
        self.assertIsNone(root.find(".//c:valAx/c:scaling/c:min", self.ns))

    def test_line_chart_keeps_auto_min(self):
        root = self._build("line", values=(3.8, 4.1, 4.5))
        self.assertIsNone(root.find(".//c:valAx/c:scaling/c:min", self.ns))

    def test_default_style_emits_value_labels_for_column(self):
        from transpiler.registries import ChartStyleRegistry
        root = self._build("column", style=ChartStyleRegistry().get())
        dlbls = root.find(".//c:barChart/c:dLbls", self.ns)
        self.assertIsNotNone(dlbls, "default style must label bars")
        self.assertEqual(dlbls.find("c:showVal", self.ns).get("val"), "1")
        self.assertEqual(dlbls.find("c:showPercent", self.ns).get("val"), "0")
        # schema order: dLbls after every ser, before axId
        plot = root.find(".//c:barChart", self.ns)
        tags = [child.tag.split("}")[1] for child in plot]
        self.assertLess(tags.index("ser"), tags.index("dLbls"))
        self.assertLess(tags.index("dLbls"), tags.index("axId"))

    def test_data_labels_none_overrides_style(self):
        from transpiler.registries import ChartStyleRegistry
        root = self._build(
            "column", style=ChartStyleRegistry().get(), data_labels="none"
        )
        self.assertIsNone(root.find(".//c:barChart/c:dLbls", self.ns))

    def test_line_chart_never_emits_labels(self):
        from transpiler.registries import ChartStyleRegistry
        root = self._build("line", style=ChartStyleRegistry().get("minimal"))
        self.assertIsNone(root.find(".//c:lineChart/c:dLbls", self.ns))

    def test_pie_default_style_emits_value_labels(self):
        from transpiler.registries import ChartStyleRegistry
        root = self._build("pie", style=ChartStyleRegistry().get())
        dlbls = root.find(".//c:pieChart/c:dLbls", self.ns)
        self.assertIsNotNone(dlbls)
        self.assertEqual(dlbls.find("c:showVal", self.ns).get("val"), "1")

    def test_invalid_data_labels_rejected_with_clear_message(self):
        with self.assertRaises(ValueError) as ctx:
            self._build("column", data_labels="banana")
        self.assertIn("banana", str(ctx.exception))
        self.assertIn("value", str(ctx.exception))

    def test_stacked_column_grouping_and_overlap(self):
        root = self._build("column", stacked="true")
        plot = root.find(".//c:barChart", self.ns)
        self.assertEqual(plot.find("c:grouping", self.ns).get("val"), "stacked")
        self.assertEqual(plot.find("c:overlap", self.ns).get("val"), "100")
        tags = [child.tag.split("}")[1] for child in plot]
        self.assertLess(tags.index("overlap"), tags.index("axId"))

    def test_percent_stacked_area_grouping(self):
        root = self._build("area", stacked="percent")
        plot = root.find(".//c:areaChart", self.ns)
        self.assertEqual(plot.find("c:grouping", self.ns).get("val"), "percentStacked")

    def test_combo_zero_baseline_only_on_bar_hosting_axis(self):
        series = [
            {"name": "rev", "values": [3.8, 4.1, 4.5], "type": "column", "axis": "primary"},
            {"name": "pct", "values": [3.1, 3.4, 3.9], "type": "line", "axis": "secondary"},
        ]
        root = self._build("combo", series_list=series)
        val_axes = root.findall(".//c:valAx", self.ns)
        self.assertEqual(len(val_axes), 2)
        by_pos = {ax.find("c:axPos", self.ns).get("val"): ax for ax in val_axes}
        self.assertIsNotNone(
            by_pos["l"].find("c:scaling/c:min", self.ns),
            "primary axis hosts columns -> zero baseline",
        )
        self.assertIsNone(
            by_pos["r"].find("c:scaling/c:min", self.ns),
            "secondary axis is line-only -> auto",
        )

    def test_chart_title_emitted_first_with_auto_title_kept(self):
        root = self._build("column", title="Run cost (£m)")
        chart = root.find(".//c:chart", self.ns)
        first = chart[0]
        self.assertEqual(first.tag.split("}")[1], "title")
        a_ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main", **self.ns}
        t = first.find("c:tx/c:rich/a:p/a:r/a:t", a_ns)
        self.assertEqual(t.text, "Run cost (£m)")
        self.assertEqual(
            chart.find("c:autoTitleDeleted", self.ns).get("val"), "0"
        )

    def test_no_title_suppresses_powerpoint_auto_title(self):
        root = self._build("column")
        chart = root.find(".//c:chart", self.ns)
        self.assertIsNone(chart.find("c:title", self.ns))
        self.assertEqual(
            chart.find("c:autoTitleDeleted", self.ns).get("val"), "1"
        )
