"""Source-material readers: hand-offs the agent can actually see.

Documents, spreadsheets, and PDFs become text/tables the agent can lift deck
content from — the engine already consumes arbitrary data through <chart> and
<table>; this closes the gap of the agent being blind to the user's files.

No new dependencies: xlsx via openpyxl (already present), docx via lxml over
the zip (we are an OOXML shop), pdf via poppler's pdftotext (already present
for rendering), csv/txt/md as text.
"""

from __future__ import annotations

import shutil
import subprocess
import zipfile
from pathlib import Path

from lxml import etree

MAX_CHARS = 24_000
MAX_SHEET_ROWS = 120

W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


class SourceReadError(ValueError):
    """Raised when a source file can't be read."""


def read_source(path: str | Path, sheet: str | None = None, max_chars: int = MAX_CHARS) -> str:
    """Read a source file into agent-consumable text (tables as pipe rows)."""
    path = Path(path)
    if not path.exists():
        raise SourceReadError(f"no such source file: {path}")
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        text = _read_xlsx(path, sheet)
    elif suffix == ".docx":
        text = _read_docx(path)
    elif suffix == ".pdf":
        text = _read_pdf(path)
    elif suffix in (".csv", ".txt", ".md", ".json", ".xml"):
        text = path.read_text(encoding="utf-8", errors="replace")
    else:
        raise SourceReadError(
            f"unsupported source type {suffix!r} (xlsx, docx, pdf, csv, txt, md, json, xml)"
        )
    if len(text) > max_chars:
        text = text[:max_chars] + f"\n... [truncated at {max_chars} chars]"
    return text


def _read_xlsx(path: Path, sheet: str | None) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    names = workbook.sheetnames
    targets = [sheet] if sheet else names
    if sheet and sheet not in names:
        raise SourceReadError(f"{path.name} has no sheet {sheet!r}; sheets: {names}")

    pieces = [f"Workbook {path.name} — sheets: {', '.join(names)}"]
    for name in targets:
        ws = workbook[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if all(value is None for value in row):
                continue
            rows.append("| " + " | ".join("" if v is None else str(v) for v in row) + " |")
            if len(rows) >= MAX_SHEET_ROWS:
                rows.append(f"| ... truncated at {MAX_SHEET_ROWS} rows |")
                break
        pieces.append(f"\n## Sheet: {name} ({len(rows)} rows shown)\n" + "\n".join(rows))
    workbook.close()
    return "\n".join(pieces)


def _read_docx(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as package:
            root = etree.fromstring(package.read("word/document.xml"))
    except (zipfile.BadZipFile, KeyError) as exc:
        raise SourceReadError(f"{path.name} is not a readable .docx") from exc

    def runs_text(element) -> str:
        return "".join(t.text or "" for t in element.iter(f"{{{W_NS}}}t"))

    pieces: list[str] = []
    body = root.find(f"{{{W_NS}}}body")
    for child in body if body is not None else []:
        tag = etree.QName(child).localname
        if tag == "p":
            text = runs_text(child)
            if text.strip():
                pieces.append(text)
        elif tag == "tbl":
            for tr in child.iter(f"{{{W_NS}}}tr"):
                cells = [runs_text(tc) for tc in tr.iter(f"{{{W_NS}}}tc")]
                pieces.append("| " + " | ".join(cells) + " |")
    return "\n".join(pieces)


def _read_pdf(path: Path) -> str:
    pdftotext = shutil.which("pdftotext")
    if not pdftotext:
        raise SourceReadError("pdftotext not on PATH (poppler-utils)")
    result = subprocess.run(
        [pdftotext, "-layout", str(path), "-"],
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        raise SourceReadError(f"pdftotext failed on {path.name}: {result.stderr[:200]}")
    return result.stdout
