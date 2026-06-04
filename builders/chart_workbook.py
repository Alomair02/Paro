#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Filename: chart_workbook.py
Description: Builds the embedded .xlsx workbook that backs a chart, so the
chart is editable in PowerPoint. The workbook values MUST match the chart's
cached values (the cache-equals-workbook invariant).
"""

import io

from openpyxl import Workbook


def build_chart_workbook_bytes(categories, series_list, chart_type="column") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    if chart_type == "scatter":
        # Per series: two columns (x, y). Series 0 -> A,B ; series 1 -> C,D ; ...
        for s_idx, series in enumerate(series_list):
            x_col = chr(ord("A") + s_idx * 2)
            y_col = chr(ord("A") + s_idx * 2 + 1)
            ws[f"{x_col}1"] = f"{series['name']} X"
            ws[f"{y_col}1"] = series["name"]
            for r_idx, (xv, yv) in enumerate(series["xy"], start=2):
                ws[f"{x_col}{r_idx}"] = xv
                ws[f"{y_col}{r_idx}"] = yv
    else:
        if not categories:
            ws["A1"] = series_list[0]["name"]   # A1 = series name (part's tx references $A$1)
            # no-category types (histogram): raw values in column A
            for r_idx, val in enumerate(series_list[0]["values"], start=2):
                ws[f"A{r_idx}"] = val
        else:
            for i, cat in enumerate(categories, start=2):
                ws[f"A{i}"] = cat
            for s_idx, series in enumerate(series_list):
                col = chr(ord("B") + s_idx)
                ws[f"{col}1"] = series["name"]
                for r_idx, val in enumerate(series["values"], start=2):
                    ws[f"{col}{r_idx}"] = val

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def build_treemap_workbook_bytes(categories, series_list):
    series = series_list[0]
    records = series["points"]                 # [{"path": [outer..inner], "value": n}, ...]
    depth = len(records[0]["path"])
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    size_col = depth + 1                        # label cols A..(depth), then size col
    ws.cell(row=1, column=size_col, value=series.get("name", "Size"))
    for i, rec in enumerate(records):
        r = i + 2
        for k, label in enumerate(rec["path"]):   # outermost-first: path[0] -> col A (k=0 -> col 1)
            ws.cell(row=r, column=k + 1, value=label)
        ws.cell(row=r, column=size_col, value=rec["value"])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()